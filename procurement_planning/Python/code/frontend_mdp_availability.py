import sys
import openpyxl
from PyQt5.QtWidgets import (
    QApplication, QMainWindow, QVBoxLayout, QWidget, QPushButton,
    QTableWidget, QTableWidgetItem, QLabel, QLineEdit, QHBoxLayout,
    QInputDialog, QDesktopWidget, QShortcut,
    QSizePolicy, QMessageBox
)
from PyQt5.QtWebEngineWidgets import QWebEngineView
from PyQt5.QtGui import QFont, QPixmap, QKeySequence
import plotly.graph_objects as go
from openpyxl.utils import get_column_letter
from backend_mdp_availability import run_gurobi_solver
from PyQt5.QtCore import Qt
import os


class MainWindow(QMainWindow):
    def __init__(self):
        super().__init__()
        # determine screen size
        screen_geometry = QDesktopWidget().availableGeometry()
        screen_width = screen_geometry.width()
        screen_height = screen_geometry.height()

        # set window size to screen size
        self.resize(screen_width, screen_height)
        self.setWindowFlags(Qt.WindowShadeButtonHint)
        self.setWindowTitle("Stochastic availability model")

        # main layout
        layout = QVBoxLayout()

        # horizontal header with logo und text
        header_layout = QHBoxLayout()

        # logo
        logo_label = QLabel()
        logo_pixmap = QPixmap("logo.png").scaled(60, 60, Qt.KeepAspectRatio)
        logo_label.setPixmap(logo_pixmap)
        header_layout.addWidget(logo_label)

        # text
        text_label = QLabel("Operations Management Group\nInstitute of Management and Economics")
        text_label.setFont(QFont("Arial", 14, QFont.Normal))
        header_layout.addWidget(text_label)

        header_layout.addStretch()

        # add the header to the main layout
        layout.addLayout(header_layout)

        # input of parameters
        self.param_inputs = {}
        self.params = {
            "d_max": 10, "x_max": 20, "y_max": 15, "pi": 5,
            "h": 1, "k": 5, "v": 20, "par_pD": 0.5, "par_pY": 0.5,
            "mu_D": 0, "sigma_D": 0, "mu_Y": 0, "sigma_Y": 0
        }

        param_layout = QHBoxLayout()
        for param, default in self.params.items():
            label = QLabel(param)
            label.setFont(QFont("Arial", 10, QFont.Bold))
            input_field = QLineEdit(str(default))
            input_field.setMaximumWidth(50)
            self.param_inputs[param] = input_field
            param_layout.addWidget(label)
            param_layout.addWidget(input_field)
        layout.addLayout(param_layout)

        # create a horizontal layout
        button_layout = QHBoxLayout()

        # add Run Solver button to the left
        self.run_button = QPushButton("Run Gurobi solver")
        self.run_button.setFont(QFont("Arial", 10, QFont.Bold))
        self.run_button.setStyleSheet("background-color: lightgreen;")
        self.run_button.clicked.connect(self.run_solver)
        self.run_button.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Fixed)

        # add Save Result button to the right
        self.save_button = QPushButton("Save results to Excel file")
        self.save_button.setFont(QFont("Arial", 10, QFont.Bold))
        self.save_button.setStyleSheet("background-color: yellow;")
        self.save_button.clicked.connect(self.save_to_excel)
        self.save_button.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Fixed)
        
        button_layout.addWidget(self.run_button)
        button_layout.addWidget(self.save_button)

        # add this layout to the main layout
        layout.addLayout(button_layout)

        # results table
        self.order_table = QTableWidget()
        self.order_table.setColumnCount(3)
        self.order_table.setHorizontalHeaderLabels(["Inventory Level", "Order Quantity ", "Probability"])
        # access the horizontal header and apply the font
        header = self.order_table.horizontalHeader()
        header.setFont(QFont("Arial", 10, QFont.Bold))
        # set all columns to have the same width
        column_width = int(screen_width/3)
        for col in range(self.order_table.columnCount()):
            self.order_table.setColumnWidth(col, column_width)
        # make the table non-editable
        self.order_table.setEditTriggers(QTableWidget.NoEditTriggers)
        layout.addWidget(self.order_table)

        self.performance_results_label = QLabel("")
        self.performance_results_label.setFont(QFont("Arial", 12))
        layout.addWidget(self.performance_results_label)

        # display plot in QWebEngineView
        self.web_view = QWebEngineView()
        self.web_view.setHtml("<h1></h1>")
        layout.addWidget(self.web_view)
        
        # central widget
        container = QWidget()
        container.setLayout(layout)
        self.setCentralWidget(container)

        # add a shortcut to close the window (Ctrl+Q)
        close_shortcut = QShortcut(QKeySequence("Ctrl+Q"), self)
        close_shortcut.activated.connect(self.close)

        self.results = None
        self.performance_results = None

    def validate_distribution_params(self, params):
        # check for conflicting distributions (binomial vs normal) for D
        if params["par_pD"] > 0.0 and (params["mu_D"] > 0.0 or params["sigma_D"] > 0.0):
            self.show_error_message("Please provide parameters for only one distribution function for D "
                                    "(binomial or normal).")
            return False

        # check for conflicting distributions (binomial vs normal) for Y
        if params["par_pY"] > 0.0 and (params["mu_Y"] > 0.0 or params["sigma_Y"] > 0.0):
            self.show_error_message("Please provide parameters for only one distribution function for Y "
                                    "(binomial or normal).")
            return False

        # check if D has invalid parameters
        if params["par_pD"] == 0.0 and params["mu_D"] == 0.0 and params["sigma_D"] == 0.0:
            self.show_error_message("Please provide parameter values for a distribution function for D.")
            return False

        # check if Y has invalid parameters
        if params["par_pY"] == 0.0 and params["mu_Y"] == 0.0 and params["sigma_Y"] == 0.0:
            self.show_error_message("Please provide parameter values for a distribution function for Y.")
            return False
        
        # check for invalid normal distribution for D
        if (params["mu_D"] > 0.0 and params["sigma_D"] == 0.0) or (params["mu_D"] == 0.0 and params["sigma_D"] > 0.0):
            self.show_error_message("Invalid normal distribution parameters for D: Both mu and sigma must be "
                                    "greater than 0.")
            return False

        # check for invalid normal distribution for Y
        if (params["mu_Y"] > 0.0 and params["sigma_Y"] == 0.0) or (params["mu_Y"] == 0.0 and params["sigma_Y"] > 0.0):
            self.show_error_message("Invalid normal distribution parameters for Y: Both mu and sigma must be "
                                    "greater than 0.")
            return False

        return True

    def run_solver(self):
        # try to convert all parameters
        params = {}
        for key, input_field in self.param_inputs.items():
            try:
                # try to convert the text in a number
                params[key] = float(input_field.text())

                # check if the value is negative
                if params[key] < 0:
                    self.show_error_message(f"The Entry for '{key}' cannot be negative. Please enter a valid number.")
                    return  # prevent the execution of the solver if the value is negative
           
            except ValueError:
                # display an error message if there is an error
                self.show_error_message(f"The Entry for '{key}' is not valid. Please enter a valid number.")
                return  # prevent the execution of the solver if an error occurred
        
        if not self.validate_distribution_params(params):
            return  # stop the execution if the validation fails
        
        self.params = params
        # execute the solver and save the results
        self.results, self.performance_results = run_gurobi_solver(params)

        # update the table
        self.update_table(self.results)

        # update the performance label
        self.update_performance_label()

        # update the diagram
        self.plot_results(self.results)

    def show_error_message(self, message):
        # show error message
        msg_box = QMessageBox()
        msg_box.setIcon(QMessageBox.Critical)
        msg_box.setText(message)
        msg_box.setWindowTitle("Error")
        msg_box.exec_()    

    def update_table(self, df):
        self.order_table.setRowCount(len(df))
        for i, row in df.iterrows():
            self.order_table.setItem(i, 0, QTableWidgetItem(str(row["Inventory Level"])))
            self.order_table.setItem(i, 1, QTableWidgetItem(str(row["Order Quantity"])))
            self.order_table.setItem(i, 2, QTableWidgetItem(f"{row['Probability']:.10f}"))
    
    def update_performance_label(self):
        result_text = "<table style=border-collapse: collapse; margin: 10px;>"
        for param, value in self.performance_results.items():
            result_text += f""" <tr>
                        <td style="padding-right: 20px; text-align: left; font-weight: bold;">{param}:</td>
                        <td style="text-align: right; padding-left: 20px; color: green;">{value:.4f}</td>
                        </tr>
                                """
        result_text += "</table>"
        self.performance_results_label.setText(result_text)

    def plot_results(self, df):
        fig = go.Figure()

        # probability vs inventory level (on secondary y-axis)
        fig.add_trace(go.Scatter(
            x=df["Inventory Level"],
            y=df["Probability"],
            mode='lines+markers',
            name='Probability',
            line=dict(color='blue'),
            marker=dict(size=6),
            yaxis='y2'  # use secondary y-axis
        ))

        # order quantity vs inventory level
        fig.add_trace(go.Scatter(
            x=df["Inventory Level"],
            y=df["Order Quantity"],
            mode='lines+markers',
            name='Order Quantity',
            line=dict(color='green'),
            marker=dict(size=6)
        ))

        # update layout with secondary y-axis
        fig.update_layout(
            title="Policy and probabilities",
            xaxis_title="Inventory Level",
            yaxis=dict(title="Order Quantity"),
            yaxis2=dict(
                title="Probability",
                overlaying='y',  # overlay y-axis 2 on y-axis 1
                side='right',    # place on the right
                showgrid=False
            ),
            template="plotly_dark",
            legend=dict(
                orientation="h",  # horizontal legend
                yanchor="bottom",
                y=1.02,
                xanchor="right",
                x=1
            )
        )

        # set the HTML into the WebEngineView
        html = fig.to_html(include_plotlyjs="cdn")
        self.web_view.setHtml(html)

    def save_to_excel(self):
        if not (os.path.isfile("results_MDP_availability.xlsx")):
            workbook = openpyxl.Workbook()
            workbook.save(filename="results_MDP_availability.xlsx")

        # get the name of the sheet
        sheet_name, ok = QInputDialog.getText(self, "Sheet-name", "Enter a sheet-name:")

        if ok and sheet_name:
            self.save_performance_results_in_excel("results_MDP_availability.xlsx", sheet_name)
    
    def save_performance_results_in_excel(self, filename, sheet_name):
      
        try:
            workbook = openpyxl.load_workbook(filename)
        except FileNotFoundError:
            workbook = openpyxl.Workbook()

        # create or choose a sheet
        if sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
        else:
            sheet = workbook.create_sheet(sheet_name)

        # create a complete list of the headers: parameters and result fields
        headers = list(self.params.keys()) + ["Expected total cost per period", "Expected inventory",
                                              "Maximum inventory", "Expected shortage", "Maximum shortage",
                                              "Expected order quantity", "Expected supply quantity"]

        # write the header to first row if the sheet is empty
        if sheet.max_row == 1 and sheet.max_column == 1 and sheet.cell(1, 1).value is None:
            for col, header in enumerate(headers, start=1):
                sheet.cell(row=1, column=col, value=header)

        # write the data
        start_row = sheet.max_row + 1
    
        # values of the parameters
        row = list(self.params.values())
        # values of the results
        row.extend([
            self.performance_results["Expected total cost per period"],
            self.performance_results["Expected inventory level"], 
            self.performance_results["Maximum inventory level"],
            self.performance_results["Expected shortage"], 
            self.performance_results["Maximum shortage"],
            self.performance_results["Expected order quantity"],
            self.performance_results["Expected supply quantity"]
        ])
        # write the row
        for col, value in enumerate(row, start=1):
            sheet.cell(row=start_row, column=col, value=value)

        # adjust the column widths
        for col_idx, col in enumerate(headers, start=1):
            col_letter = get_column_letter(col_idx)
            max_length = max(
                len(str(sheet.cell(row=row, column=col_idx).value)) for row in range(1, sheet.max_row + 1)
            )
            adjusted_width = max_length + 2  # Padding
            sheet.column_dimensions[col_letter].width = adjusted_width

        # save the Excel file
        try:
            # try to save the workbook
            workbook.save(filename)
            QMessageBox.information(None, "Success", f"The file has been successfully saved as '{filename}'.")
        except PermissionError:
            # handle the case where the file is open or permission is denied
            QMessageBox.critical(None, "File Error",
                                 f"Cannot save to '{filename}' because the file is open or you lack permission. "
                                 "Please close the file and try again.")
        except Exception as e:
            # catch other unexpected errors and display the error message
            QMessageBox.critical(None, "Unexpected Error",
                                 f"An error occurred while saving the file: {e}")


if __name__ == "__main__":
    app = QApplication(sys.argv)
    main_window = MainWindow()
    main_window.show()
    sys.exit(app.exec_())
