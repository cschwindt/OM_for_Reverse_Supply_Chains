import sys
import openpyxl
from PyQt5.QtWidgets import (
    QApplication, QMainWindow, QVBoxLayout, QWidget, QPushButton,
    QTableWidget, QTableWidgetItem, QLabel, QLineEdit, QHBoxLayout, QFileDialog,QInputDialog,QDesktopWidget,QShortcut,
    QSizePolicy, QFileDialog, QMessageBox
)
from PyQt5.QtWebEngineWidgets import QWebEngineView
from PyQt5.QtGui import QFont , QPixmap ,QKeySequence
import plotly.graph_objects as go
from openpyxl.utils import get_column_letter
from mdp_policy_avaibility import run_gurobi_solver
from PyQt5.QtCore import Qt

class MainWindow(QMainWindow):
    def __init__(self):
        super().__init__()
                # Bildschirmgröße ermitteln
        screen_geometry = QDesktopWidget().availableGeometry()
        screen_width = screen_geometry.width()
        screen_height = screen_geometry.height()

        # Fenstergröße auf Bildschirmgröße setzen
        self.resize(screen_width, screen_height)
        self.setWindowFlags(Qt.FramelessWindowHint)

        # Hauptlayout
        layout = QVBoxLayout()

        # Horizontaler Header mit Logo und Text
        header_layout = QHBoxLayout()

        # Logo
        logo_label = QLabel()
        logo_pixmap = QPixmap("Beschaffungsplaung\Python\documents\logo.png").scaled(60, 60, Qt.KeepAspectRatio)  # Logo skalieren
        logo_label.setPixmap(logo_pixmap)
        header_layout.addWidget(logo_label)

        # Text
        text_label = QLabel("Operations Management Group\nInstitute of Management and Economics")
        text_label.setFont(QFont("Arial", 14, QFont.Normal))
        header_layout.addWidget(text_label)
    
        # Platzhalter, um Header linksbündig zu halten
        header_layout.addStretch()

        # Header ins Hauptlayout
        layout.addLayout(header_layout)

        # Parameter-Eingabe
        self.param_inputs = {}
        self.params = {
            "dmax": 10, "xmax": 20, "ymax": 15, "pi": 5,
            "h": 1, "k": 0, "v": 20, "par_pD": 0, "par_pY": 0 ,
            "mu_D" : 0 , "sigma_D" : 0, "mu_Y" : 0 , "sigma_Y" : 0
        }

        param_layout = QHBoxLayout()
        for param, default in self.params.items():
            label = QLabel(param)
            label.setFont(QFont("Georgia", 10, QFont.Bold))
            input_field = QLineEdit(str(default))
            self.param_inputs[param] = input_field
            param_layout.addWidget(label)
            param_layout.addWidget(input_field)
        layout.addLayout(param_layout)

            # Create a horizontal layout
        button_layout = QHBoxLayout()

        # Add the "run solver" button to the left
        self.run_button = QPushButton("run gurobi-solver")
        self.run_button.setFont(QFont("Arial", 10, QFont.Bold))
        self.run_button.setStyleSheet("background-color: lightgreen;")
        self.run_button.clicked.connect(self.run_solver)
        self.run_button.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Fixed)

        # Add the "save result" button to the right
        self.save_button = QPushButton("save result as excel file")
        self.save_button.setFont(QFont("Arial", 10, QFont.Bold))
        self.save_button.setStyleSheet("background-color: yellow;")
        self.save_button.clicked.connect(self.save_to_excel)
        self.save_button.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Fixed)
        
        button_layout.addWidget(self.run_button)
        button_layout.addWidget(self.save_button)

        # Add this layout to the main layout
        layout.addLayout(button_layout)

        # Ergebnisse Tabelle 
        self.order_table = QTableWidget()
        self.order_table.setColumnCount(3)
        self.order_table.setHorizontalHeaderLabels(["Inventory Level", "Order Quantity ", "Probability"])
        # Access the horizontal header and apply the font
        header = self.order_table.horizontalHeader()
        header.setFont(QFont("Arial",10, QFont.Bold))
        # Set all columns to have the same width
        column_width = int(screen_width/3)
        for col in range(self.order_table.columnCount()):
            self.order_table.setColumnWidth(col, column_width)
        # Make the table non-editable
        self.order_table.setEditTriggers(QTableWidget.NoEditTriggers)
        layout.addWidget(self.order_table)

        self.performance_results_label = QLabel("")
        self.performance_results_label.setFont(QFont("Arial", 12))
        layout.addWidget(self.performance_results_label)

        # Plot im QWebEngineView anzeigen
        self.web_view = QWebEngineView()
        self.web_view.setHtml("<h1></h1>")
        layout.addWidget(self.web_view)
        
        # Zentrales Widget
        container = QWidget()
        container.setLayout(layout)
        self.setCentralWidget(container)

        # Add a shortcut to close the window (Ctrl+Q)
        close_shortcut = QShortcut(QKeySequence("Ctrl+Q"), self)
        close_shortcut.activated.connect(self.close)

        # Datenplatzhalter
        self.results = None
        self.performance_results = None

    def validate_distribution_params(self, params):
        # Check for conflicting distributions (binomial vs normal) for D
        if params["par_pD"] > 0.0 and (params["mu_D"] > 0.0 or params["sigma_D"] > 0.0):
            self.show_error_message("Please provide parameters for only one distribution function for D (binomial or normal).")
            return False

        # Check for conflicting distributions (binomial vs normal) for Y
        if params["par_pY"] > 0.0 and (params["mu_Y"] > 0.0 or params["sigma_Y"] > 0.0):
            self.show_error_message("Please provide parameters for only one distribution function for Y (binomial or normal).")
            return False

        # Check if D has no valid parameters
        if params["par_pD"] == 0.0 and params["mu_D"] == 0.0 and params["sigma_D"] == 0.0:
            self.show_error_message("Please provide parameter values for a distribution function for D.")
            return False

        # Check if Y has no valid parameters
        if params["par_pY"] == 0.0 and params["mu_Y"] == 0.0 and params["sigma_Y"] == 0.0:
            self.show_error_message("Please provide parameter values for a distribution function for Y.")
            return False
        
        # Check for invalid normal distribution for D
        if (params["mu_D"] > 0.0 and params["sigma_D"] == 0.0) or (params["mu_D"] == 0.0 and params["sigma_D"] > 0.0):
            self.show_error_message("Invalid normal distribution parameters for D: Both mu and sigma must be greater than 0.")
            return False

        # Check for invalid normal distribution for Y
        if (params["mu_Y"] > 0.0 and params["sigma_Y"] == 0.0) or (params["mu_Y"] == 0.0 and params["sigma_Y"] > 0.0):
            self.show_error_message("Invalid normal distribution parameters for Y: Both mu and sigma must be greater than 0.")
            return False


        return True


    def run_solver(self):
        # Versuche, alle Parameter-Werte zu konvertieren
        params = {}
        for key, input_field in self.param_inputs.items():
            try:
                # Versuche den Text in eine Zahl umzuwandeln
                params[key] = float(input_field.text())

                # Überprüfe, ob der Wert negativ ist
                if params[key] < 0:
                    self.show_error_message(f"The Entry for '{key}' cannot be negative. Please enter a valid number.")
                    return  # Verhindere die Ausführung des Solvers, wenn der Wert negativ ist
           
            except ValueError:
                # Falls es einen Fehler gibt (keine gültige Zahl), zeige ein Fehler-Popup
                self.show_error_message(f"The Entry for '{key}' is not valid. Please enter a valid number.")
                return  # Verhindere die Ausführung des Solvers, wenn Fehler aufgetreten ist
        
        if not self.validate_distribution_params(params):
              return  # Stop execution if validation fails
        
        self.params = params
        # Solver aufrufen und Ergebnisse speichern
        self.results , self.performance_results = run_gurobi_solver(params)

        # Tabelle aktualisieren
        self.update_table(self.results)

        # Performance label update
        self.update_performance_label()

        # Grafik aktualisieren
        self.plot_results(self.results)

    def show_error_message(self, message):
        # Zeige eine Fehlermeldung an
        msg_box = QMessageBox()
        msg_box.setIcon(QMessageBox.Critical)
        msg_box.setText(message)
        msg_box.setWindowTitle("Fehler")
        msg_box.exec_()    

    def update_table(self, df):
        self.order_table.setRowCount(len(df))
        for i, row in df.iterrows():
            self.order_table.setItem(i, 0, QTableWidgetItem(str(row["Inventory Level"])))
            self.order_table.setItem(i, 1, QTableWidgetItem(str(row["Order Quantity"])))
            self.order_table.setItem(i, 2, QTableWidgetItem(f"{row['Probability']:.10f}"))
    
    def update_performance_label(self):
         result_text = "<table style=border-collapse: collapse; margin: 10px;>"
         for param, value in self.performance_results.items():
            result_text += f""" <tr>
                        <td style="padding-right: 20px; text-align: left; font-weight: bold;">{param}:</td>
                        <td style="text-align: right; padding-left: 20px; color: green;">{value:.4f}</td>
                        </tr>
                                """
         result_text += "</table>"  
         self.performance_results_label.setText(result_text)

    def plot_results(self, df):
        fig = go.Figure()

        # Probability vs Inventory Level (on secondary y-axis)
        fig.add_trace(go.Scatter(
            x=df["Inventory Level"],
            y=df["Probability"],
            mode='lines+markers',
            name='Probability',
            line=dict(color='blue'),
            marker=dict(size=6),
            yaxis='y2'  # Use secondary y-axis
        ))

        # Order Quantity vs Inventory Level
        fig.add_trace(go.Scatter(
            x=df["Inventory Level"],
            y=df["Order Quantity"],
            mode='lines+markers',
            name='Order Quantity',
            line=dict(color='green'),
            marker=dict(size=6)
        ))

        # Update layout with secondary y-axis
        fig.update_layout(
            title="Inventory Level Analysis",
            xaxis_title="Inventory Level",
            yaxis=dict(title="Order Quantity"),
            yaxis2=dict(
                title="Probability",
                overlaying='y',  # Overlay y-axis 2 on y-axis 1
                side='right',    # Place on the right
                showgrid=False
            ),
            template="plotly_dark",
            legend=dict(
                orientation="h",  # Horizontal legend
                yanchor="bottom",
                y=1.02,
                xanchor="right",
                x=1
            )
        )

        # Set the HTML into the WebEngineView
        html = fig.to_html(include_plotlyjs="cdn")
        self.web_view.setHtml(html)

    def save_to_excel(self):
        # Wählen Sie eine bestehende Excel-Datei aus
        file_path, _ = QFileDialog.getOpenFileName(self, "select excel-file ", "", "excel-file (*.xlsx)")

        if not file_path:
            return

        # Sheet-Name abfragen
        sheet_name, ok = QInputDialog.getText(self, "Sheet-name", "Enter a sheet-name:")

        if ok and sheet_name:
            self.save_performance_results_in_excel(file_path, sheet_name)
    
    def save_performance_results_in_excel(self, filename, sheet_name):
      
        try:
            workbook = openpyxl.load_workbook(filename)
        except FileNotFoundError:
            workbook = openpyxl.Workbook()

        # Sheet erstellen oder auswählen
        if sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
        else:
            sheet = workbook.create_sheet(sheet_name)

        # Erstelle die vollständige Liste der Header: Parameter + Ergebnis-Felder
        headers = list(self.params.keys()) + ["Expected total cost per period", "Expected Inventory", "Maximum Inventory",
            "Expected Shortage", "Maximum Shortage", "Expected Order"
        ]

        # Schreibe Header in die erste Zeile, falls das Sheet leer ist
        if sheet.max_row == 1 and sheet.max_column == 1 and sheet.cell(1, 1).value is None:
            for col, header in enumerate(headers, start=1):
                sheet.cell(row=1, column=col, value=header)

        # Schreibe die Daten
        start_row = sheet.max_row + 1
    
        # Werte der Parameter
        row = list(self.params.values())
        # Werte der Ergebnisse
        row.extend([
            self.performance_results["Expected total cost per period"],
            self.performance_results["Expected inventory level"], 
            self.performance_results["Maximum inventory level"],
            self.performance_results["Expected shortage"], 
            self.performance_results["Maximum shortage"],
            self.performance_results["Expected order quantity"]
        ])
        # Schreibe die Zeile
        for col, value in enumerate(row, start=1):
            sheet.cell(row=start_row, column=col, value=value)

        # Spaltenbreiten anpassen
        for col_idx, col in enumerate(headers, start=1):
            col_letter = get_column_letter(col_idx)
            max_length = max(
                len(str(sheet.cell(row=row, column=col_idx).value)) for row in range(1, sheet.max_row + 1)
            )
            adjusted_width = max_length + 2  # Padding
            sheet.column_dimensions[col_letter].width = adjusted_width

        # Speichere die Excel-Datei
        try:
            # Attempt to save the workbook
            workbook.save(filename)
            QMessageBox.information(None, "Success", f"The file has been successfully saved as '{filename}'.")
        except PermissionError:
            # Handle the case where the file is open or permission is denied
            QMessageBox.critical(None, "File Error",
                                f"Cannot save to '{filename}' because the file is open or you lack permission. "
                                "Please close the file and try again.")
        except Exception as e:
            # Catch other unexpected errors and display the error message
            QMessageBox.critical(None, "Unexpected Error",
                                f"An error occurred while saving the file: {e}")     

if __name__ == "__main__":
    app = QApplication(sys.argv)
    main_window = MainWindow()
    main_window.show()
    sys.exit(app.exec_())
